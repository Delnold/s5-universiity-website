from django.core.handlers.wsgi import WSGIRequest
from django.db.models import QuerySet
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from main_website.models import Champions
from S5_universiity.settings import MEDIA_URL

def excel_view(request: WSGIRequest, champions: QuerySet = Champions.objects.all()):
    wb: Workbook = Workbook()
    ws: Workbook.active = wb.active

    header = ['Champion_Name', 'Champion_Story', 'Champion_Icon_URL']
    ws.append(header)
    champ_url = None
    for champ in champions:
        champ_url = r"http://" + request.get_host() + MEDIA_URL + str(champ.champion_icon)
        ws.append([str(champ.champion_name), str(champ.champion_story), champ_url])


    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Champions_excel.xlsx'
    wb.save(response)

    return response

def excel_admin_upload(excel_file):
    try:
        workbook = load_workbook(excel_file, read_only=True)
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            Champion_Name, Champion_Story = row
            champion = Champions(champion_name=Champion_Name, champion_story=Champion_Story)
            champion.save()
    except Exception as e:
        return False